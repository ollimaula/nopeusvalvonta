import win32com.client as win32 # type: ignore
from xlsxwriter import Workbook
from tkinter import Button, Canvas, Label, Tk, IntVar, END, Entry, messagebox, Frame, N, E, S, ttk
from datetime import datetime
from getpass import getuser
from string import capwords
from os import makedirs
from os.path import expanduser, isfile
from openpyxl import load_workbook

def quit_program():
    if messagebox.askokcancel('Lopeta', 'Haluatko varmasti poistua?\nTallentamattomat tiedot menetetään.'):
        GUI.root.quit()
        GUI.root.destroy()

class data_lists: # Listat datan tallentamiseksi.

    timelist = [] # syötteen aloitusajat
    speedlist = [] # käyttäjän syöttämät nopeudet
    reglist = [] # käyttäjän syöttämät rekisterinumerot
    namelist = [] # käyttäjän syöttämät kuljettajat
    companylist = [] # käyttäjän syöttämät yritykset
    light_speeding_list = [] # nopeudet, jotka täyttävät lievän ylinopeuden määritelmän
    heavy_speeding_list = [] # nopeudet, jotka täyttävät vakavan ylinopeuden määritelmän
    host_companies = (
        'forcit', 
        'vihtavuori', 
        'nammo vihtavuori', 
        'navi', 
        'nala', 
        'nammo lapua', 
        'lapua'
    ) # lista yritysten nimistä; oikeaan muotoon korjattavaksi.

class excel_file:

        home = expanduser('~') # polku käyttäjän kotikansioon
        current_time = datetime.now() # Tallennetaan tämänhetkinen päivä ja aika
        file_date = current_time.strftime('%d.%m.%Y') # Tallennetaan tämänhetkinen päivämäärä tiedoston nimeä varten.
        file_name = f'Nopeusvalvonta {file_date}.xlsx'
        file_location = f'{home}\\Nopeusvalvonta\\' # polku tiedostojen tallentamista varten

        file = f'{file_location}{file_name}'

class get_data: # Pyydetään käyttäjältä dataa ja käsitellään sitä tarpeen mukaan.

    speed = 1000 # Annetaan speed_input:n loopille alkutilanne
    reg_num = None
    driver_name = None
    company = None

    def __init__(self):
        self.data_collection()

    def speed_check(self, speed):

        # Tarkastetaan, ylittääkö käyttäjän syöttämä nopeus ylinopeudeksi määritellyt rajat.

        if speed > 30 and speed <= 35:
            data_lists.light_speeding_list.append(f'{speed} km/h')
            data_lists.light_speeding_list.append(self.reg_num)
            data_lists.light_speeding_list.append(self.driver_name)
            data_lists.light_speeding_list.append(self.company_check(self.company))

        elif speed > 35:
            data_lists.heavy_speeding_list.append(f'{speed} km/h')
            data_lists.heavy_speeding_list.append(self.reg_num)
            data_lists.heavy_speeding_list.append(self.driver_name)
            data_lists.heavy_speeding_list.append(self.company_check(self.company))

        else:
            pass

        self.speed = f'{speed} km/h'
        return data_lists.speedlist.append(self.speed)

    def company_check(self, company): 

        # Tarkastetaan, löytyykö käyttäjän määrittelemä yritys etukäteen määriteltyjen yritysten joukosta.

        company_name=company

        if company.lower() in data_lists.host_companies:

            if company.lower() in (data_lists.host_companies[1], data_lists.host_companies[2], data_lists.host_companies[3]):
                company_name='Nammo Vihtavuori Oy'
                return company_name

            elif company.lower() in (data_lists.host_companies[4], data_lists.host_companies[5], data_lists.host_companies[6]):
                company_name='Nammo Lapua Oy'
                return company_name

            elif company.lower() in data_lists.host_companies[0]:
                company_name='Forcit Oy'
                return company_name

        else:
            return company_name

    def speed_input(self):

        back_button = Button(GUI.wrapper_frame, text='Edellinen', state='disabled', bg='#2B2D2F', fg='#D3D3D3') # placeholder button
        back_button.grid(row=2, column=1)

        speed_error_prompt = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Syöttämäsi tieto on virheellinen, yritä uudestaan.')
        speed_error_prompt_2 = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Anna nopeus väliltä 4 - 999 km/h.')
        speed_prompt = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Mitattu nopeus:')
        speed_prompt.grid(row=1, column=0)

        while self.speed >= 1000 or self.speed <= 0:

            try:

                GUI.entry_text.delete(0, END)

                check = IntVar()

                enter_button = Button(GUI.wrapper_frame, text='Hyväksy (Enter)', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=lambda: check.set(1))
                enter_button.grid(row=2, column=2)
                GUI.root.bind('<Return>', lambda event: check.set(1))

                enter_button.wait_variable(check) # Odotetaan, kunnes käyttäjä painaa valitsemaansa painiketta.

                speed = int(GUI.entry_text.get())
                self.speed = speed-3

                if speed_error_prompt.winfo_ismapped():
                    speed_error_prompt.grid_forget()
                speed_error_prompt_2.grid(row=3, column=0)

            except ValueError:
                if speed_error_prompt_2.winfo_ismapped():
                    speed_error_prompt_2.grid_forget()
                speed_error_prompt.grid(row=3, column=0)

        speed_error_prompt.destroy()
        speed_error_prompt_2.destroy()
        speed_prompt.destroy()
        self.regnum_input()

    def regnum_input(self):

        regnum_error_prompt = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Syöttämäsi tieto on virheellinen, yritä uudestaan.\nSyötä rekisterinumero muodossa \'ABC-123\' tai \'abc-123\'.')
        regnum_prompt = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Ajoneuvon rekisterinumero:')
        regnum_prompt.grid(row=1, column=0)

        while self.reg_num == None: # Jatketaan kunnes reg_num saa kelvollisen arvon.

            try:
                GUI.entry_text.delete(0, END)

                check = IntVar()
                back = IntVar()

                enter_button = Button(GUI.wrapper_frame, text='Hyväksy (Enter)', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=lambda: check.set(1))
                enter_button.grid(row=2, column=2)
                GUI.root.bind('<Return>', lambda event: check.set(1))

                back_button = Button(GUI.wrapper_frame, text='Edellinen', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=lambda: [back.set(1), check.set(1)])
                back_button.grid(row=2, column=1)

                enter_button.wait_variable(check) # Odotetaan, kunnes käyttäjä painaa valitsemaansa painiketta.

                if back.get() == 1:
                    self.speed = 1000
                    return back_button.destroy(), regnum_error_prompt.destroy(), regnum_prompt.destroy(), self.speed_input()

                reg_num_for_checking=(GUI.entry_text.get())

                if len(reg_num_for_checking) == 7 and reg_num_for_checking[0:2].isalpha() and reg_num_for_checking[3] == '-' and reg_num_for_checking[4:6].isnumeric(): # Varmistetaan, että käyttäjän syöttämä rekisterinumero on kelvollinen.
                    self.reg_num = reg_num_for_checking.upper()

                elif len(reg_num_for_checking) == 6 and reg_num_for_checking[0:2].isalpha() and reg_num_for_checking[3:5].isnumeric(): # Tarkastetaan onko rekkari muuten ok, mutta väliviiva puuttuu.
                    reg_num_repaired = f'{reg_num_for_checking[0:3]}-{reg_num_for_checking[3:6]}' # Lisätään väliviiva.
                    self.reg_num = reg_num_repaired.upper()

                else:
                    regnum_error_prompt.grid(row=3, column=0)
                
            except ValueError:
                regnum_error_prompt.grid(row=3, column=0)
        
        regnum_error_prompt.destroy()
        regnum_prompt.destroy()
        back_button.destroy()
        self.driver_input()

    def driver_input(self):

        back_button = Button(GUI.wrapper_frame, text='Edellinen', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=self.regnum_input)
        back_button.grid(row=2, column=1)

        driver_prompt = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Kuljettajan nimi:')
        driver_prompt.grid(row=1, column=0)

        driver_lower_prompt = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Jätä kuljettajan nimi tyhjäksi, mikäli se ei ole tiedossasi.')
        driver_lower_prompt.grid(row=3, column=0)

        while self.driver_name == None or len(self.driver_name) > 35: # jatketaan looppia, kunnes käyttäjä antaa tyhjän tai maksimissaan 25 merkkiä olevan nimen.

            GUI.entry_text.delete(0, END)

            check = IntVar()
            back = IntVar()

            enter_button = Button(GUI.wrapper_frame, text='Hyväksy (Enter)', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=lambda: check.set(1))
            enter_button.grid(row=2, column=2)
            GUI.root.bind('<Return>', lambda event: check.set(1))

            back_button = Button(GUI.wrapper_frame, text='Edellinen', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=lambda: [back.set(1), check.set(1)])
            back_button.grid(row=2, column=1)

            enter_button.wait_variable(check) # Odotetaan, kunnes käyttäjä painaa valitsemaansa painiketta.

            if back.get() == 1:
                self.reg_num = None
                return back_button.destroy(), driver_prompt.destroy(), driver_lower_prompt.destroy(), self.regnum_input()
            
            driver_name = GUI.entry_text.get()
            self.driver_name = capwords(driver_name, sep=None)
            
            driver_lower_prompt.destroy()
            driver_lower_prompt = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Jätä kuljettajan nimi tyhjäksi, mikäli se ei ole tiedossasi.\nKuljettajan nimen maksimipituus on 35 kirjainta.')
            driver_lower_prompt.grid(row=3, column=0)

        driver_prompt.destroy()
        driver_lower_prompt.destroy()
        back_button.destroy()
        self.company_input()

    def company_input(self):

        back_button=Button(GUI.wrapper_frame, text='Edellinen', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=self.driver_input)
        back_button.grid(row=2, column=1)

        company_prompt = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Kuljettajan yritys:')
        company_prompt.grid(row=1, column=0)

        company_lower_prompt = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Jätä yrityksen nimi tyhjäksi, mikäli se ei ole tiedossasi.')
        company_lower_prompt.grid(row=3, column=0)

        while self.company == None or len(self.company) > 25: # jatketaan looppia, kunnes käyttäjä antaa tyhjän tai maksimissaan 25 merkkiä olevan nimen.

            GUI.entry_text.delete(0, END)

            check = IntVar()
            back = IntVar()

            enter_button = Button(GUI.wrapper_frame, text='Hyväksy (Enter)', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=lambda: check.set(1))
            enter_button.grid(row=2, column=2)
            GUI.root.bind('<Return>', lambda event: check.set(1))

            back_button = Button(GUI.wrapper_frame, text='Edellinen', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=lambda: [back.set(1), check.set(1)])
            back_button.grid(row=2, column=1)

            enter_button.wait_variable(check) # Odotetaan, kunnes käyttäjä painaa valitsemaansa painiketta.

            if back.get() == 1:
                self.driver_name = None
                return back_button.destroy(), company_prompt.destroy(), company_lower_prompt.destroy(), self.driver_input()
            
            company = GUI.entry_text.get()
            self.company = capwords(company, sep=None)

            company_lower_prompt.destroy()
            company_lower_prompt = Label(GUI.wrapper_frame, bg='#1B1D1F', fg='#D3D3D3', text='Jätä yrityksen nimi tyhjäksi, mikäli se ei ole tiedossasi.\nYrityksen nimen maksimipituus on 25 kirjainta.')
            company_lower_prompt.grid(row=3, column=0) 

        company_prompt.destroy()
        company_lower_prompt.destroy()
        back_button.destroy()

    def data_collection(self):

        current_time = datetime.now() # Tallennetaan tämän hetkinen päivä ja aika
        data_time = current_time.strftime('%H:%M') # Tallennetaan tämän hetkinen aika muodossa HH:MM

        GUI.root.unbind('<Tab>')

        GUI.status_bar_prompt.destroy()
        GUI.status_bar_prompt = Label(GUI.status_bar, font=('', 8), bg='#252729', fg='#D3D3D3', text=f'Tallennetaan uutta dataa.')
        GUI.status_bar_prompt.pack(side='left')

        GUI.entry_text.config(state='normal')
        GUI.entry_text.focus_set()

        GUI.data_button.config(state='disabled', cursor='')

        self.speed_input()

        GUI.entry_text.delete(0, END)
        GUI.entry_text.config(state='disabled')

        GUI.data_button.config(state='normal', cursor='hand2')

        fake_button = Button(GUI.wrapper_frame, text='Hyväksy (Enter)', state='disabled', bg='#2B2D2F', fg='#D3D3D3') # placeholder button
        fake_button.grid(row=2, column=2)

        back_button = Button(GUI.wrapper_frame, text='Edellinen', state='disabled', bg='#2B2D2F', fg='#D3D3D3') # placeholder button
        back_button.grid(row=2, column=1)
        
        GUI.root.bind('<Tab>', lambda event:get_data())

        GUI.status_bar_prompt.destroy()
        GUI.status_bar_prompt = Label(GUI.status_bar, font=('', 8), bg='#252729', fg='#D3D3D3', text=f'Tallennettu uusi tieto.')
        GUI.status_bar_prompt.pack(side='left')
        
        data_lists.timelist.append(data_time)
        self.speed_check(self.speed)
        data_lists.reglist.append(self.reg_num)
        data_lists.namelist.append(self.driver_name)
        data_lists.companylist.append(self.company_check(self.company))
        data_handling.show_data(self)

class data_handling: # Käsitellään käyttäjältä kerättyä dataa.

    def __init__(self):

        if isfile(excel_file.file):
            if messagebox.askyesno('Nopeusvalvonta', 'Tälle päivämäärälle löytyy nopeusvalvonnan tiedosto.\nHaluatko jatkaa tietojen tallentamista kyseiseen tiedostoon?'):
                self.pull_data()

        self.show_data()       

    def pull_data(self):

        existing_workbook = load_workbook(excel_file.file)
        sheet = existing_workbook.active

        for row in sheet.iter_cols(min_col=2, max_col=2):
            for cell in row[2:]:
                if cell.value is not None:
                    data_lists.timelist.append(cell.value)
                    
        for row in sheet.iter_cols(min_col=3, max_col=3):
            for cell in row[2:]:
                if cell.value is not None:
                    data_lists.speedlist.append(cell.value)
        
        for row in sheet.iter_cols(min_col=4, max_col=4):
            for cell in row[2:]:
                if cell.value is not None:
                    data_lists.reglist.append(cell.value)

        for row in sheet.iter_cols(min_col=5, max_col=5):
            for cell in row[2:]:
                if cell.value is not None:
                    data_lists.namelist.append(cell.value)

        for row in sheet.iter_cols(min_col=6, max_col=6):
            for cell in row[2:]:
                if cell.value is not None:
                    data_lists.companylist.append(cell.value)

        row_counter = 4

        for row in sheet.iter_rows(min_row=3, max_col=13):
            row_counter += 1
            for cell in row[9:]:
                if cell.value is not None:
                    data_lists.light_speeding_list.append(cell.value)
            if cell.value is None:
                break

        for row in sheet.iter_rows(min_row=row_counter, max_col=13):
            for cell in row[9:]:
                if cell.value is not None:
                    data_lists.heavy_speeding_list.append(cell.value)

    def show_data(self):

        # näytetään tallennettu data omassa framessaan.

        scroll_style = ttk.Style() # muokataan ttk.scrollbarin tyyliä
        scroll_style.theme_use('clam')
        scroll_style.layout( # luodaan scrollbarille nuoleton teema.
            'Vertical.TScrollbar',
            [('Vertical.Scrollbar.trough',
            {'children': [('Vertical.Scrollbar.thumb',
            {'expand': '1', 'sticky': 'nswe'})],
            'sticky': 'ns'})]
        )
        scroll_style.map( # muokataan scrollbarin värejä ja ulkoasua.
            'Vertical.TScrollbar',
            background=[ ('!active','#2B2D2F'),('pressed', '#1B1D1F'), ('active', '#252729')]
        )
        scroll_style.configure( # muokataan scrollbarin värejä ja ulkoasua.
            'Vertical.TScrollbar', 
            gripcount=0, 
            darkcolor='#1B1D1F', 
            lightcolor='gray',
            troughcolor='#D3D3D3',
        )

        scrollframe_style = ttk.Style() # muokataan ttk.framen tyyliä
        scrollframe_style.configure('TFrame', background='black')

        inner_data_frame = Frame(GUI.data_frame, width=585, height=280, bg='black')
        inner_data_frame.grid(row=0, column=0)
        inner_data_frame.grid_propagate(0)

        data_canvas = Canvas(inner_data_frame, width=568, height=280, bg='black', borderwidth=0, highlightthickness=0)
        data_canvas.grid(row=0, column=0)

        data_frame_scrollable = ttk.Frame(data_canvas)
        data_frame_scrollable.grid(row=0, column=0)

        title_label_1 = Label(data_frame_scrollable, bg='black', fg='#D3D3D3', text='AIKA').grid(row=0, column=0)
        title_label_2 = Label(data_frame_scrollable, bg='black', fg='#D3D3D3', text='NOPEUS').grid(row=0, column=1)
        title_label_3 = Label(data_frame_scrollable, bg='black', fg='#D3D3D3', text='REK.NRO').grid(row=0, column=2)
        title_label_4 = Label(data_frame_scrollable, bg='black', fg='#D3D3D3', text='KULJETTAJA').grid(row=0, column=3)
        title_label_5 = Label(data_frame_scrollable, bg='black', fg='#D3D3D3', text='YRITYS').grid(row=0, column=4)

        # käännetään listat ympäri, jotta tuorein data on ylimpänä ruudulla.

        time_label = Label(data_frame_scrollable, bg='black', fg='#D3D3D3', text='\n'.join(data_lists.timelist.__reversed__()))
        time_label.grid(row=1, column=0, padx=3)

        speed_label = Label(data_frame_scrollable, bg='black', fg='#D3D3D3', text='\n'.join(data_lists.speedlist.__reversed__()))
        speed_label.grid(row=1, column=1, padx=3)

        reg_label = Label(data_frame_scrollable, bg='black', fg='#D3D3D3', text='\n'.join(data_lists.reglist.__reversed__()))
        reg_label.grid(row=1, column=2, padx=3)

        name_label = Label(data_frame_scrollable, bg='black', fg='#D3D3D3', text='\n'.join(data_lists.namelist.__reversed__()))
        name_label.grid(row=1, column=3, padx=3)

        company_label = Label(data_frame_scrollable, bg='black', fg='#D3D3D3', text='\n'.join(data_lists.companylist.__reversed__()))
        company_label.grid(row=1, column=4, padx=3)

        # bindataan mousewheel scrollbariin, mikäli kursori on canvasin/scrollbarin päällä.

        def using_mousewheel(event): # käyttäjä liikuttaa rullaa
            data_canvas.yview_scroll(int(-1*(event.delta/120)), 'units')
        def bind_to_mousewheel(event): # kursori widgetin päällä
            data_canvas.bind_all('<MouseWheel>', using_mousewheel)
        def unbind_from_mousewheel(event): # kursori EI widgetin päällä
            data_canvas.unbind_all('<MouseWheel>')

        if len(data_lists.timelist) >= 18: # sallitaan scroll vasta kun on tarve.
            data_scrollbar = ttk.Scrollbar(inner_data_frame, orient='vertical', command=data_canvas.yview)
            data_scrollbar.grid(row=0, column=1, sticky=N+E+S)

            data_canvas.create_window((0, 0), window=data_frame_scrollable, anchor='nw')
            data_canvas.configure(yscrollcommand=data_scrollbar.set)

            data_canvas.bind('<Enter>', bind_to_mousewheel)
            data_canvas.bind('<Leave>', unbind_from_mousewheel)
            data_scrollbar.bind('<Enter>', bind_to_mousewheel)
            data_scrollbar.bind('<Leave>', unbind_from_mousewheel)

        data_frame_scrollable.bind('<Configure>', lambda e: data_canvas.configure(scrollregion=data_canvas.bbox('all'))) # päivitetään scrollattava alue.

    def create_workbook():

        def progress():
            GUI.progress_bar['value'] += 10
            GUI.root.update_idletasks()

        def auto_fit(): # Autofit workbookin soluille

            try:

                try:
                    excel = win32.gencache.EnsureDispatch('Excel.Application')

                except AttributeError:
                    # jos -> 'module' object has no attribute 'CLSIDToPackageMap'
                    # tyhjennetään cache ja koitetaan uudestaan
                    import os
                    import re
                    import sys
                    import shutil
                    emergency_modules = [m.__name__ for m in sys.modules.values()]
                    for module in emergency_modules:
                        if re.match(r'win32com\.gen_py\..+', module):
                            del sys.modules[module]
                    shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
                    excel = win32.gencache.EnsureDispatch('Excel.Application')

                progress()
                wb = excel.Workbooks.Open(excel_file.file)
                ws = wb.Worksheets('Sheet1')
                ws.Columns.AutoFit()
                wb.Save()
                excel.Application.Quit()
                progress()

            except Exception:
                messagebox.showerror('Excel error', 'Virhe excel tiedoston käsittelyssä.\nVarmista, että tietokoneellasi on ajantasainen Excel ohjelmisto.')

        if isfile(f'{excel_file.file}'):
            if messagebox.askokcancel('File exists', 'Olet tallentamassa vanhan tiedoston päälle.\nOletko aivan varma tästä?'):
                pass
            else:
                return

        try:
            makedirs(excel_file.file_location, exist_ok = True) # Luodaan polkuun tarvittavat kansiot, mikäli tarpeen.

            workbook = Workbook(excel_file.file) # Luodaan excel workbook.
            worksheet = workbook.add_worksheet() # Avataan uusi worksheet.

            timerow = 2
            speedrow = 2
            light_speeding_row = 2
            regrow = 2
            namerow = 2
            companyrow = 2
            column = 1

            title_format = workbook.add_format({'bold': True})
            
            worksheet.write('B2', 'Aika', title_format)
            worksheet.write('C2', 'Ajonopeus', title_format)
            worksheet.write('D2', 'Rekisterinumero', title_format)
            worksheet.write('E2', 'Kuljettajan nimi', title_format)
            worksheet.write('F2', 'Kuljettajan yritys', title_format)
            worksheet.write('J2', 'Lievät ylinopeudet', title_format)
            progress()

            for item in data_lists.timelist:
                worksheet.write(timerow, column, item)
                timerow +=1
            column +=1
            progress()

            for item in data_lists.speedlist:
                worksheet.write(speedrow, column, item)
                speedrow +=1
            column +=1
            progress()

            for item in data_lists.reglist:
                worksheet.write(regrow, column, item)
                regrow +=1
            column +=1
            progress()

            for item in data_lists.namelist:
                worksheet.write(namerow, column, item)
                namerow +=1
            column +=1
            progress()

            for item in data_lists.companylist:
                worksheet.write(companyrow, column, item)
                companyrow +=1
            column +=1
            progress()

            speedingcolumn = column +3

            for item in data_lists.light_speeding_list: # Rakennetaan excel lista lievistä ylinopeuksista
                worksheet.write(light_speeding_row, speedingcolumn, item)
                speedingcolumn +=1
                if speedingcolumn == column +7:
                    speedingcolumn = column +3
                    light_speeding_row +=1
            progress()

            heavy_speeding_row = light_speeding_row +2

            worksheet.write(heavy_speeding_row -1, speedingcolumn, 'Vakavat ylinopeudet', title_format)
        
            for item in data_lists.heavy_speeding_list: # Rakennetaan excel lista vakavista ylinopeuksista
                worksheet.write(heavy_speeding_row, speedingcolumn, item)
                speedingcolumn +=1
                if speedingcolumn == column +7:
                    speedingcolumn = column +3
                    heavy_speeding_row +=1
            progress()
                
            workbook.close()
            auto_fit()

            GUI.progress_bar['value'] = 0

            GUI.status_bar_prompt.destroy()
            GUI.status_bar_prompt = Label(GUI.status_bar, font=('', 8), bg='#252729', fg='#D3D3D3', text=f'Tiedot tallennettu: C:\\Users\\{getuser()}\\Nopeusvalvonta')
            GUI.status_bar_prompt.pack(side='left')

        except Exception: 
            messagebox.showerror('Permission Error', 'Ei oikeutta tiedoston käsittelyyn.\nVarmista, ettei tiedosto ole auki.')
            GUI.status_bar_prompt.destroy()
            GUI.status_bar_prompt = Label(GUI.status_bar, font=('', 8), bg='#252729', fg='#D3D3D3', text=f'Tietojen tallentaminen epäonnistui.')
            GUI.status_bar_prompt.pack(side='left')
            GUI.progress_bar['value'] = 0

class GUI: # Graafinen käyttöliittymä.

    def __init__(self):
        self.draw_gui()

    def draw_gui(self):
        self.wrapper_frame.pack()
        self.data_button.grid(row=0, column=0, pady=10)
        self.workbook_button.grid(row=0, column=1, pady=10)
        self.quit_button.grid(row=0, column=2, pady=10)
        self.space_reserver.grid(row=1, column=0)
        self.enter_button.grid(row=2, column=2)
        self.back_button.grid(row=2, column=1)
        self.entry_text.grid(row=2, column=0)
        self.space_reserver_2.grid(row=3, column=0)
        self.space_reserver_3.grid(row=4, column=0)
        self.status_bar.pack(side='bottom', fill='x')
        self.status_bar_border.pack(side='bottom', fill='x')
        self.progress_bar.pack(side='right', padx=2)

    # root config
    root = Tk()
    root.title('Nopeusvalvonta')
    root.minsize(750, 600)
    root.config(bg='#1B1D1F')

    # wrapperi guille
    wrapper_frame = Frame(root, bg='#1B1D1F', pady=40)

    # alue tallennetulle datalle
    data_frame = Frame(wrapper_frame, bg='black', relief='sunken', width=600, height=300, borderwidth=10)
    data_frame.grid(row=5, column=0, columnspan=3)
    data_frame.grid_propagate(0)

    # painikkeet
    data_button = Button(wrapper_frame, text='Rekisteröi uusi mitattu nopeus (Tab)', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=get_data)
    root.bind('<Tab>', lambda event:get_data())

    workbook_button = Button(wrapper_frame, text='Tallenna tiedot taulukkoon', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=data_handling.create_workbook)
    quit_button = Button(wrapper_frame, text='Lopeta', cursor='hand2', bg='#2B2D2F', fg='#D3D3D3', command=quit_program)
    enter_button = Button(wrapper_frame, text='Hyväksy (Enter)', bg='#2B2D2F', fg='#D3D3D3', state='disabled') # placeholder button
    back_button = Button(wrapper_frame, text='Edellinen', bg='#2B2D2F', fg='#D3D3D3', state='disabled') # placeholder button

    # input kenttä
    entry_text = Entry(wrapper_frame, width=50, bg='black', fg='#D3D3D3', insertbackground='#D3D3D3', borderwidth=5, disabledbackground='#2B2D2F',  state='disabled') # Luodaan inputille tekstikenttä

    # tyhjiä labeleitä tilan varaamiseksi
    space_reserver = Label(wrapper_frame, bg='#1B1D1F', text='') # Varataan gridistä paikkoja. Ei ehkä paras/oikea tapa tehdä tämä, mutta if it works it works.
    space_reserver_2 = Label(wrapper_frame, bg='#1B1D1F', text='\n')
    space_reserver_3 = Label(wrapper_frame, bg='#1B1D1F', text='')
    status_bar_prompt = Label()

    # status bar
    status_bar = Frame(root, relief='sunken', height=20, bg='#252729')
    status_bar_border = Frame(root, height=1, bg='#16181a')

    progressbar_style = ttk.Style() # muokataan ttk.progressbarin tyyliä
    progressbar_style.theme_use('clam')
    progressbar_style.configure(
        'Horizontal.TProgressbar', 
        background='#33FF33', 
        troughcolor='#1B1D1F',
        bordercolor='#252729',
        darkcolor='#1B1D1F',
        lightcolor='grey'
    )
    progress_bar = ttk.Progressbar(status_bar, orient='horizontal', length=200, mode='determinate')

def main():
    GUI()
    GUI.root.after_idle(data_handling) # after_idle, koska messagebox data_handlingin initissä rikkoo muuten event loopin
    GUI.root.mainloop()

if __name__ == '__main__':
    main()